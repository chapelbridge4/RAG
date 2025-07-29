from typing import Dict, Any, List, Optional
import asyncio
from pathlib import Path
import mimetypes
import hashlib

# Document processing imports
try:
    import pypdf
except ImportError:
    pypdf = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

class DocumentProcessor:
    """Advanced document processor for multiple file formats"""
    
    def __init__(self):
        self.supported_formats = {
            'text/plain': self._process_text,
            'application/pdf': self._process_pdf,
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': self._process_docx,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': self._process_xlsx,
            'text/markdown': self._process_text,
            'text/html': self._process_html,
        }
    
    async def process_file(self, file_content: bytes, filename: str, metadata: Dict[str, Any] = None) -> Dict[str, Any]:
        """Process uploaded file and extract text content"""
        try:
            # Determine file type
            mime_type, _ = mimetypes.guess_type(filename)
            if not mime_type:
                mime_type = 'text/plain'
            
            # Generate content hash for deduplication
            content_hash = hashlib.md5(file_content).hexdigest()
            
            # Extract text based on file type
            if mime_type in self.supported_formats:
                text_content = await self.supported_formats[mime_type](file_content, filename)
            else:
                # Fallback to treating as text
                text_content = file_content.decode('utf-8', errors='ignore')
            
            return {
                'filename': filename,
                'content': text_content,
                'content_hash': content_hash,
                'mime_type': mime_type,
                'size': len(file_content),
                'metadata': metadata or {},
                'processing_status': 'success'
            }
            
        except Exception as e:
            return {
                'filename': filename,
                'content': '',
                'error': str(e),
                'processing_status': 'failed'
            }
    
    async def _process_text(self, content: bytes, filename: str) -> str:
        """Process plain text files"""
        return content.decode('utf-8', errors='ignore')
    
    async def _process_pdf(self, content: bytes, filename: str) -> str:
        """Process PDF files"""
        if pypdf is None:
            raise ImportError("pypdf package required for PDF processing")
        
        try:
            from io import BytesIO
            pdf_reader = pypdf.PdfReader(BytesIO(content))
            
            text_content = []
            for page in pdf_reader.pages:
                text_content.append(page.extract_text())
            
            return '\n\n'.join(text_content)
        except Exception as e:
            raise Exception(f"Failed to process PDF: {e}")
    
    async def _process_docx(self, content: bytes, filename: str) -> str:
        """Process DOCX files"""
        if DocxDocument is None:
            raise ImportError("python-docx package required for DOCX processing")
        
        try:
            from io import BytesIO
            doc = DocxDocument(BytesIO(content))
            
            text_content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content.append(paragraph.text)
            
            return '\n\n'.join(text_content)
        except Exception as e:
            raise Exception(f"Failed to process DOCX: {e}")
    
    async def _process_xlsx(self, content: bytes, filename: str) -> str:
        """Process Excel files"""
        if openpyxl is None:
            raise ImportError("openpyxl package required for Excel processing")
        
        try:
            from io import BytesIO
            workbook = openpyxl.load_workbook(BytesIO(content))
            
            text_content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_content = [f"Sheet: {sheet_name}"]
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        sheet_content.append('\t'.join(row_text))
                
                text_content.append('\n'.join(sheet_content))
            
            return '\n\n'.join(text_content)
        except Exception as e:
            raise Exception(f"Failed to process Excel: {e}")
    
    async def _process_html(self, content: bytes, filename: str) -> str:
        """Process HTML files"""
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            # Fallback to simple text extraction
            html_content = content.decode('utf-8', errors='ignore')
            # Simple HTML tag removal
            import re
            clean_text = re.sub(r'<[^>]+>', ' ', html_content)
            return re.sub(r'\s+', ' ', clean_text).strip()
        
        try:
            html_content = content.decode('utf-8', errors='ignore')
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Extract text
            text = soup.get_text()
            
            # Clean up whitespace
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = ' '.join(chunk for chunk in chunks if chunk)
            
            return text
        except Exception as e:
            raise Exception(f"Failed to process HTML: {e}")
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported MIME types"""
        return list(self.supported_formats.keys())
    
    async def validate_file(self, filename: str, file_size: int, max_size: int = 10 * 1024 * 1024) -> Dict[str, Any]:
        """Validate uploaded file"""
        mime_type, _ = mimetypes.guess_type(filename)
        
        validation_result = {
            'valid': True,
            'errors': [],
            'warnings': []
        }
        
        # Check file size
        if file_size > max_size:
            validation_result['valid'] = False
            validation_result['errors'].append(f"File size ({file_size} bytes) exceeds maximum ({max_size} bytes)")
        
        # Check file type
        if mime_type and mime_type not in self.supported_formats:
            validation_result['warnings'].append(f"File type {mime_type} may not be optimally processed")
        
        # Check filename
        if not filename or len(filename) > 255:
            validation_result['valid'] = False
            validation_result['errors'].append("Invalid filename")
        
        return validation_result